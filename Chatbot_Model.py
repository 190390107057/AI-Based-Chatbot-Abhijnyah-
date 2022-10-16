import math
import numpy as np
from creme import compose
from creme import feature_extraction
from creme import naive_bayes
import pandas as pd
import speech_recognition 
import pyttsx3 as pp 
from xlwt import Workbook
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('200.xlsx')
ws = wb['Sheet1']

data = pd.read_excel('200.xlsx')
print(len(data))



df = data.iloc[:,0:2]

from sklearn.model_selection import train_test_split
message_train,message_test=train_test_split(df)

messages_train = message_train.to_records(index=False)
messages_test=message_test.to_records(index=False)

model = compose.Pipeline(
    ('tokenize', feature_extraction.BagOfWords(lowercase=True)),
    ('nb', naive_bayes.MultinomialNB(alpha=1))
)


from creme import metrics
metric=metrics.Accuracy()



# Training the model row by row
for tag, pattern in messages_train:
    # print(tag)
    # print(Pattern)
    model = model.fit_one(pattern,tag)
    # print(model)
    y_pred = model.predict_one(pattern)
    # print(y_pred)
    metric = metric.update(tag, y_pred)
print(metric)

test_metric=metrics.Accuracy()
for tag, Pattern in messages_test:
    y_pred = model.predict_one(pattern)
    test_metric = metric.update(tag, y_pred)

print(test_metric)
engine = pp.init()

voices = engine.getProperty('voices')
print(voices)

engine.setProperty('voice', voices[1].id)


def speak(word):
    engine.say(word)
    engine.runAndWait()

settings = {
            "help_email": "fakeEmail@notArealEmail.com",
            "faq_page": "https://www.aicte-india.org/search/google/faqs#gsc.tab=0&gsc.q=faqs&gsc.sort="
        }

def response(msg):
    print(msg)
    global pred
    proba = model.predict_proba_one(msg)
    maximum = max(proba.keys(), key = (lambda new_k : proba[new_k]))
    print(maximum)
    #print(maximum)
    print(proba[maximum])       
    if (proba[maximum]>= 0.1):
        pred = model.predict_one(msg)
        print(pred)
    
        for index in range(len(data['tag'])):
            if data['tag'][index] == pred:
                ans = data['responses'][index]
                print(ans)
                # return ans
 
        count = 0
        print(len(data))
        for index in range(len(data)):
            if data['pattern'][index] == msg:
                count = 1
                break
        print(count)
        if(count != 1):
            data.loc[len(data.index)] = [pred, msg, ans, 1]
            model.fit_one(msg, pred)                                                                    
            data.to_excel('demo.xlsx')
        return ans

    else:
        ans = "I'm having trouble understanding what you are saying. At the time, my ability is quite limited, " \
                  "please refer to %s or email %s if I was not able to answer your question. " \
                  "For convenience, a google link has been generated below: \n%s\n" % (settings['faq_page'],
                                                                                    settings['help_email'], "https://www.aicte-india.org/search/google/%s" %
                                                                                    ("+".join(msg.split(" "))))  
    print(ans)                                                                     
    return ans
    

"""def audio_to_Text():
    sr = speech_recognition.Reccognizer()
    with speech_recognition.Microphone()as m:
        sr.adjust_for_ambient_noise(m, duration=0.2) 
        audio = sr.listen(m)
        query = sr.recognize_google(audio, language='eng-in')
        query = query.capitalize() """


    


